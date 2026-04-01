import io
import os
import platform
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping

# --- Функция для поиска системного шрифта ---
def get_system_font_path():
    system = platform.system()
    possible_paths = []
    if system == "Windows":
        possible_paths = [
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/times.ttf",
            "C:/Windows/Fonts/arialbd.ttf",
            "C:/Windows/Fonts/timesbd.ttf"
        ]
    elif system == "Darwin":  # macOS
        possible_paths = [
            "/System/Library/Fonts/Arial.ttf",
            "/System/Library/Fonts/Times.ttf",
            "/System/Library/Fonts/Arial Bold.ttf",
            "/System/Library/Fonts/Times Bold.ttf"
        ]
    elif system == "Linux":
        possible_paths = [
            "/usr/share/fonts/truetype/msttcorefonts/Arial.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        ]
    for path in possible_paths:
        if os.path.exists(path):
            return path
    project_dejavu = os.path.join(os.path.dirname(__file__), 'DejaVuSans.ttf')
    if os.path.exists(project_dejavu):
        return project_dejavu
    return None

def get_bold_font_path(regular_path):
    if not regular_path:
        return None
    dirname = os.path.dirname(regular_path)
    basename = os.path.basename(regular_path)
    name, ext = os.path.splitext(basename)
    bold_name = name.replace('Regular', 'Bold').replace('Sans', 'Sans-Bold')
    bold_path = os.path.join(dirname, bold_name + ext)
    if os.path.exists(bold_path):
        return bold_path
    if 'arial' in basename.lower():
        bold_path = os.path.join(dirname, 'arialbd.ttf')
        if os.path.exists(bold_path):
            return bold_path
    if 'times' in basename.lower():
        bold_path = os.path.join(dirname, 'timesbd.ttf')
        if os.path.exists(bold_path):
            return bold_path
    return None

PDF_FONT_NAME = 'Helvetica'
font_path = get_system_font_path()
if font_path:
    try:
        pdfmetrics.registerFont(TTFont('CyrillicFont', font_path))
        PDF_FONT_NAME = 'CyrillicFont'
        bold_path = get_bold_font_path(font_path)
        if bold_path:
            pdfmetrics.registerFont(TTFont('CyrillicFont-Bold', bold_path))
            addMapping('CyrillicFont', 0, 0, 'CyrillicFont')
            addMapping('CyrillicFont', 1, 0, 'CyrillicFont-Bold')
    except Exception as e:
        print(f"Не удалось загрузить шрифт {font_path}: {e}")

def sanitize_sheet_title(title):
    invalid_chars = r'[]:*?/\\'
    for ch in invalid_chars:
        title = title.replace(ch, '_')
    if len(title) > 31:
        title = title[:31]
    return title

def generate_excel_report(data, group_code, start_date, end_date,
                          total_lessons, total_students,
                          total_present, total_late, total_excused, total_absent,
                          total_possible, max_possible):
    """
    data: список студентов с ключами:
        full_name, subgroup, present, late, excused, absent,
        total, total_hours, attended, attended_hours, percent
    """
    wb = Workbook()
    sheet_title = sanitize_sheet_title(f"Отчет {group_code}")
    ws = wb.active
    ws.title = sheet_title

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    total_hours = total_lessons * 2
    max_hours = total_possible * 2
    attended_hours = (total_present + total_late) * 2
    absent_hours = (total_absent + total_excused) * 2
    attendance_rate = (attended_hours / max_hours * 100) if max_hours else 0

    # Заголовок
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1, value=f"Отчет о посещаемости")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:G2')
    group_cell = ws.cell(row=2, column=1, value=f"Группа: {group_code}")
    group_cell.font = Font(size=12)
    group_cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:G3')
    period_cell = ws.cell(row=3, column=1, value=f"Период: {start_date} - {end_date}")
    period_cell.font = Font(size=12)
    period_cell.alignment = Alignment(horizontal='center')

    # Общая статистика (основные показатели)
    ws.cell(row=5, column=1, value="Всего пар:").font = Font(bold=True)
    ws.cell(row=5, column=2, value=total_lessons)
    ws.cell(row=6, column=1, value="Всего часов:").font = Font(bold=True)
    ws.cell(row=6, column=2, value=total_hours)

    ws.cell(row=8, column=1, value="Всего студентов:").font = Font(bold=True)
    ws.cell(row=8, column=2, value=total_students)
    ws.cell(row=9, column=1, value="Всего человеко-часов:").font = Font(bold=True)
    ws.cell(row=9, column=2, value=max_hours)
    ws.cell(row=10, column=1, value="Фактически посещено (человеко-часов):").font = Font(bold=True)
    ws.cell(row=10, column=2, value=attended_hours)
    ws.cell(row=11, column=1, value="Пропущено (человеко-часов):").font = Font(bold=True)
    ws.cell(row=11, column=2, value=absent_hours)

    # Таблица абсолютных значений (в парах)
    stats_headers = ["", "Присутствовало", "Опоздало", "По уважительной", "Без уважительной"]
    stats_values = ["Всего (в парах)", total_present, total_late, total_excused, total_absent]
    start_row = 13
    for col, header in enumerate(stats_headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    for col, value in enumerate(stats_values, start=1):
        cell = ws.cell(row=start_row+1, column=col, value=value)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Анализ посещаемости (проценты)
    ws.cell(row=start_row+3, column=1, value="Анализ посещаемости").font = Font(bold=True, size=12)
    ws.cell(row=start_row+4, column=1, value="Общая посещаемость:").font = Font(bold=True)
    ws.cell(row=start_row+4, column=2, value=f"{attendance_rate:.1f}%")
    ws.cell(row=start_row+5, column=1, value="Посещено (человеко-часов):").font = Font(bold=True)
    ws.cell(row=start_row+5, column=2, value=f"{attended_hours} ({attendance_rate:.1f}%)")
    ws.cell(row=start_row+6, column=1, value="Пропущено (человеко-часов):").font = Font(bold=True)
    ws.cell(row=start_row+6, column=2, value=f"{absent_hours} ({100 - attendance_rate:.1f}%)")
    ws.cell(row=start_row+7, column=1, value="Из них по уважительной (человеко-часов):").font = Font(bold=True)
    ws.cell(row=start_row+7, column=2, value=total_excused * 2)

    # Детализация по студентам (упрощённая)
    ws.cell(row=start_row+9, column=1, value="Детализация по студентам").font = Font(bold=True, size=12)

    detail_headers = [
        "Студент", "Всего пар", "Посещено пар", "Пропущено", "Пропущено по уваж.", "Опоздания", "% посещ."
    ]
    for col, header in enumerate(detail_headers, start=1):
        cell = ws.cell(row=start_row+10, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row = start_row + 11
    for student in data:
        ws.cell(row=row, column=1, value=student['full_name']).border = border
        ws.cell(row=row, column=2, value=student['total']).border = border
        ws.cell(row=row, column=3, value=student['attended']).border = border
        ws.cell(row=row, column=4, value=student['absent']).border = border
        ws.cell(row=row, column=5, value=student['excused']).border = border
        ws.cell(row=row, column=6, value=student['late']).border = border
        ws.cell(row=row, column=7, value=student['percent']).border = border
        row += 1

    # Автоширина колонок
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_len + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf_report(data, group_code, start_date, end_date,
                        total_lessons, total_students,
                        total_present, total_late, total_excused, total_absent,
                        total_possible, max_possible):
    """Генерация отчета в PDF."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=36, leftMargin=36,
                            topMargin=36, bottomMargin=36)

    # Стили
    styles = getSampleStyleSheet()
    if PDF_FONT_NAME != 'Helvetica':
        styles.add(ParagraphStyle(name='CyrillicNormal', fontName=PDF_FONT_NAME, fontSize=12, leading=14))
        styles.add(ParagraphStyle(name='CyrillicAnalysis', fontName=PDF_FONT_NAME, fontSize=14, leading=16))
        styles.add(ParagraphStyle(name='CyrillicHeader', fontName=PDF_FONT_NAME, fontSize=18, leading=22, fontWeight='bold'))
        styles.add(ParagraphStyle(name='CyrillicHeading', fontName=PDF_FONT_NAME, fontSize=16, leading=18, spaceAfter=8, fontWeight='bold'))
        styles.add(ParagraphStyle(name='CyrillicBold', fontName=PDF_FONT_NAME, fontSize=14, leading=16, fontWeight='bold'))
        styles.add(ParagraphStyle(name='DateStyle', fontName=PDF_FONT_NAME, fontSize=10, leading=12, alignment=2, textColor=colors.gray))
    else:
        styles.add(ParagraphStyle(name='CyrillicNormal', fontName='Helvetica', fontSize=12, leading=14))
        styles.add(ParagraphStyle(name='CyrillicAnalysis', fontName='Helvetica', fontSize=14, leading=16))
        styles.add(ParagraphStyle(name='CyrillicHeader', fontName='Helvetica', fontSize=18, leading=22, fontWeight='bold'))
        styles.add(ParagraphStyle(name='CyrillicHeading', fontName='Helvetica', fontSize=16, leading=18, spaceAfter=8, fontWeight='bold'))
        styles.add(ParagraphStyle(name='CyrillicBold', fontName='Helvetica', fontSize=14, leading=16, fontWeight='bold'))
        styles.add(ParagraphStyle(name='DateStyle', fontName='Helvetica', fontSize=10, leading=12, alignment=2, textColor=colors.gray))

    total_hours = total_lessons * 2
    max_hours = total_possible * 2
    attended_hours = (total_present + total_late) * 2
    absent_hours = (total_absent + total_excused) * 2
    attendance_rate = (attended_hours / max_hours * 100) if max_hours else 0

    story = []

    # Шапка
    story.append(Paragraph(f"Группа: {group_code}", styles['CyrillicHeader']))
    story.append(Paragraph(f"Период: {start_date} - {end_date}", styles['CyrillicHeader']))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"Всего пар: {total_lessons}", styles['CyrillicHeader']))
    story.append(Paragraph(f"Всего часов: {total_hours}", styles['CyrillicHeader']))
    story.append(Paragraph(f"Всего студентов: {total_students}", styles['CyrillicHeader']))
    story.append(Paragraph(f"Всего человеко-часов: {max_hours}", styles['CyrillicHeader']))
    story.append(Paragraph(f"Фактически посещено (человеко-часов): {attended_hours}", styles['CyrillicHeader']))
    story.append(Paragraph(f"Пропущено (человеко-часов): {absent_hours}", styles['CyrillicHeader']))
    story.append(Spacer(1, 16))

    # Общая статистика (абсолютные значения в парах)
    story.append(Paragraph("## Общая статистика", styles['CyrillicHeading']))
    stats_data = [
        ["Присутствовало (пар)", str(total_present)],
        ["Опоздало (пар)", str(total_late)],
        ["По уважительной (пар)", str(total_excused)],
        ["Без уважительной (пар)", str(total_absent)],
    ]
    stats_table = Table(stats_data, colWidths=[doc.width * 0.4, doc.width * 0.6])
    stats_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), PDF_FONT_NAME),
        ('FONTSIZE', (0,0), (-1,-1), 12),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,0), PDF_FONT_NAME),
        ('FONTWEIGHT', (0,0), (-1,0), 'BOLD'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 16))

    # Анализ посещаемости
    story.append(Paragraph("## Анализ посещаемости", styles['CyrillicHeading']))
    analysis_text = f"Общая посещаемость: {attendance_rate:.1f}%"
    story.append(Paragraph(f"☒ {analysis_text}", styles['CyrillicAnalysis']))
    story.append(Spacer(1, 8))
    story.append(Paragraph("Детали:", styles['CyrillicBold']))
    story.append(Paragraph(f"• Посещено (человеко-часов): {attended_hours} ({attendance_rate:.1f}%)", styles['CyrillicAnalysis']))
    story.append(Paragraph(f"• Пропущено (человеко-часов): {absent_hours} ({100 - attendance_rate:.1f}%)", styles['CyrillicAnalysis']))
    story.append(Paragraph(f"• Из них по уважительной (человеко-часов): {total_excused * 2}", styles['CyrillicAnalysis']))
    story.append(Spacer(1, 16))

    # Детализация по студентам (упрощённая)
    story.append(Paragraph("## Детализация по студентам", styles['CyrillicHeading']))
    detail_data = [[
        "Студент", "Всего пар", "Посещено", "Пропущено", "По уваж.", "Опоздания", "%"
    ]]
    for s in data:
        detail_data.append([
            s['full_name'],
            str(s['total']),
            str(s['attended']),
            str(s['absent']),
            str(s['excused']),
            str(s['late']),
            f"{s['percent']}%"
        ])

    col_widths = [
        doc.width * 0.25,  # студент
        doc.width * 0.12,  # всего пар
        doc.width * 0.15,  # посещено пар
        doc.width * 0.15,  # пропущено
        doc.width * 0.10,  # пропущено по уваж.
        doc.width * 0.13,  # опоздания
        doc.width * 0.08,  # %
    ]
    detail_table = Table(detail_data, colWidths=col_widths)
    detail_style = [
        ('FONTNAME', (0,0), (-1,-1), PDF_FONT_NAME),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,0), PDF_FONT_NAME),
        ('FONTWEIGHT', (0,0), (-1,0), 'BOLD'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
    ]
    for i in range(1, len(detail_data)):
        if i % 2 == 0:
            detail_style.append(('BACKGROUND', (0,i), (-1,i), colors.whitesmoke))
    detail_table.setStyle(TableStyle(detail_style))
    story.append(detail_table)

    # Дата формирования
    current_datetime = datetime.now().strftime("%d.%m.%Y %H:%M")
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"Отчет сформирован: {current_datetime}", styles['DateStyle']))

    doc.build(story)
    buffer.seek(0)
    return buffer